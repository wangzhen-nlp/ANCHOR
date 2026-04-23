
"""
筛选Incident Ticket记录：
1. 行的内容包含至少两个站点ID
2. 至少两个匹配站点包含指定场景对应的设备
"""
import argparse
import json
import os
import re
import warnings
from collections import defaultdict, deque

if __package__ in (None, ""):
    from _script_env import ensure_repo_root

    ensure_repo_root(1)

import pandas as pd
from alarm_tools.progress_utils import ProgressBar
from topology_resources import NE_GRAPH_JSON, SITE_DEVICE_COUNTS_JSON, resource_display
from ticket_resources import (
    DEFAULT_INCIDENT_TICKET_XLSX,
    resource_display as ticket_resource_display,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - fallback for environments without openpyxl
    load_workbook = None


DATETIME_TEXT_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}")
DATETIME_FULLMATCH_PATTERN = re.compile(r"^\s*\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\s*$")
TICKET_DATE_PATTERN = re.compile(r"-(\d{8})-")
FALLBACK_TICKET_DATE_PATTERN = re.compile(r"(?<!\d)(\d{8})(?!\d)")

SCENARIO_DEVICE_DOMAINS = {
    "transmission": ("TRANSMISSION", "Transmission"),
    "data": ("DATA", "Data"),
}


def _get_scenario_device_domain(scenario: str) -> tuple:
    normalized_scenario = str(scenario or "transmission").strip().lower()
    if normalized_scenario not in SCENARIO_DEVICE_DOMAINS:
        supported = ", ".join(sorted(SCENARIO_DEVICE_DOMAINS))
        raise ValueError(f"不支持的场景: {scenario}; 可选值: {supported}")
    return SCENARIO_DEVICE_DOMAINS[normalized_scenario]


def _empty_filter_stats() -> dict:
    return {'total': 0, 'valid': 0, 'only_one_site': 0, 'missing_required_device': 0}


def load_site_device_mapping(json_file: str) -> dict:
    """加载站点设备映射，返回站点ID -> 设备类型集合的映射"""
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 转换为：站点ID -> 设备类型集合
    site_devices = {}
    for site_id, devices in data.items():
        normalized_site_id = str(site_id).strip().upper()
        if not normalized_site_id:
            continue
        if isinstance(devices, dict):
            site_devices[normalized_site_id] = {
                str(device_type).strip().upper()
                for device_type in devices.keys()
                if str(device_type).strip()
            }
        else:
            site_devices[normalized_site_id] = set()
    return site_devices


def _normalize_match_value(value) -> str:
    return str(value).strip().upper() if value is not None else ""


def build_keyword_matcher(keyword_to_outputs: dict) -> dict:
    """基于关键词全集构建 Aho-Corasick 自动机。"""
    root = {"next": {}, "fail": None, "outputs": []}

    for keyword, outputs in keyword_to_outputs.items():
        normalized_keyword = _normalize_match_value(keyword)
        if not normalized_keyword:
            continue
        node = root
        for char in normalized_keyword:
            node = node["next"].setdefault(char, {"next": {}, "fail": None, "outputs": []})
        for output in outputs:
            if output:
                node["outputs"].append((len(normalized_keyword), output))

    queue = deque()
    for child in root["next"].values():
        child["fail"] = root
        queue.append(child)

    while queue:
        node = queue.popleft()
        for char, child in node["next"].items():
            fail = node["fail"]
            while fail is not None and char not in fail["next"]:
                fail = fail["fail"]
            child["fail"] = fail["next"][char] if fail is not None and char in fail["next"] else root
            child["outputs"].extend(child["fail"]["outputs"])
            queue.append(child)

    return root


def build_site_matcher(known_site_ids: list) -> dict:
    """基于站点全集构建 Aho-Corasick 自动机，便于按文本一次扫描匹配站点。"""
    keyword_to_outputs = {
        _normalize_match_value(site_id): [_normalize_match_value(site_id)]
        for site_id in known_site_ids
        if _normalize_match_value(site_id)
    }
    return build_keyword_matcher(keyword_to_outputs)


def load_device_site_mapping(ne_graph_file: str) -> dict:
    """加载设备到站点的映射，仅使用设备ID作为匹配关键词。"""
    with open(ne_graph_file, 'r', encoding='utf-8') as f:
        ne_graph = json.load(f)

    device_to_sites = defaultdict(set)
    if not isinstance(ne_graph, dict):
        return {}

    for ne_id, ne_info in ne_graph.items():
        if not isinstance(ne_info, dict):
            continue
        site_id = _normalize_match_value(ne_info.get('site_id', ''))
        if not site_id:
            continue
        normalized_token = _normalize_match_value(ne_id)
        if normalized_token:
            device_to_sites[normalized_token].add(site_id)

    return {
        token: sorted(site_ids)
        for token, site_ids in device_to_sites.items()
        if site_ids
    }


def build_device_matcher(device_site_mapping: dict) -> dict:
    """基于设备关键词构建设备匹配自动机，输出命中的设备ID。"""
    keyword_to_outputs = {
        device_id: [device_id]
        for device_id in device_site_mapping.keys()
        if _normalize_match_value(device_id)
    }
    return build_keyword_matcher(keyword_to_outputs)


def _row_to_text(row_values) -> str:
    return ' '.join(str(value) for value in row_values)


def extract_outputs_from_text(text: str, matcher: dict) -> list:
    """从文本中提取关键词关联输出（去重并按首次出现排序）。"""
    if not text:
        return []

    normalized_text = _normalize_match_value(text)
    first_positions = {}
    state = matcher

    for idx, char in enumerate(normalized_text):
        while state is not matcher and char not in state["next"]:
            state = state["fail"]

        if char in state["next"]:
            state = state["next"][char]
        else:
            state = matcher

        for token_len, output in state["outputs"]:
            start = idx - token_len + 1
            if output not in first_positions:
                first_positions[output] = start

    positions = sorted((pos, output) for output, pos in first_positions.items())
    return [output for pos, output in positions]


def extract_site_ids_from_text(text: str, site_matcher: dict) -> list:
    """从文本中提取所有站点ID（基于 Aho-Corasick 自动机，去重并按首次出现排序）"""
    return extract_outputs_from_text(text, site_matcher)


def check_site_devices(site_ids: list, site_device_mapping: dict, required_device_domain: str) -> tuple:
    """
    检查站点列表是否满足条件：
    - 至少2个站点
    - 至少2个站点包含指定设备域
    返回: (是否满足条件, 涉及的设备类型, 包含指定设备域的站点)
    """
    if len(site_ids) < 2:
        return False, set(), []

    all_devices = set()
    required_device_domain = str(required_device_domain or "").strip().upper()
    required_domain_sites = []

    for site_id in site_ids:
        if site_id not in site_device_mapping:
            continue

        devices = site_device_mapping[site_id]
        all_devices.update(devices)

        if required_device_domain not in devices:
            continue
        required_domain_sites.append(site_id)

    return len(required_domain_sites) >= 2, all_devices, required_domain_sites


def build_known_site_ids(site_device_mapping: dict) -> list:
    """构建稳定的站点匹配列表，长站点优先，减少短串误匹配。"""
    return sorted(site_device_mapping.keys(), key=lambda site_id: (-len(site_id), site_id))


def _print_section(title: str):
    print(f"\n[{title}]")


def _print_key_values(items):
    if not items:
        return
    width = max(len(label) for label, _ in items)
    for label, value in items:
        print(f"{label:<{width}} : {value}")


def _print_file_start(input_file: str, total_rows: int = None):
    _print_section("处理文件")
    _print_key_values([
        ("文件路径", input_file),
        ("文件名称", os.path.basename(input_file)),
    ])
    if total_rows is not None:
        _print_key_values([("原始记录数", total_rows)])
    print("开始逐行筛选...")


def _print_stats(title: str, stats: dict, scenario_device_label: str):
    _print_section(title)
    _print_key_values([
        ("总记录数", stats['total']),
        ("命中记录数", stats['valid']),
        ("不足 2 个站点", stats['only_one_site']),
        (f"具备 {scenario_device_label} 的站点不足 2 个", stats['missing_required_device']),
    ])


def _print_output_result(row_count: int, output_file: str = None, json_output_file: str = None, aggregated: bool = False):
    action = "汇总输出" if aggregated else "输出"
    items = [(f"{action}记录数", row_count)]
    if output_file:
        items.append((f"{action}Excel", output_file))
    if json_output_file:
        items.append((f"{action}JSON", json_output_file))
    _print_key_values(items)


def _count_total_sites(matched_sites_by_row) -> int:
    total = 0
    for row_sites in matched_sites_by_row or []:
        for site_id in row_sites:
            if _normalize_match_value(site_id):
                total += 1
    return total


def _normalize_ticket_key(value) -> str:
    return str(value).strip().upper() if value is not None else ""


def _derive_match_cache_output_path(output_file: str) -> str:
    base, _ = os.path.splitext(output_file)
    return f"{base}.match_cache.json"


def _load_match_cache(match_cache_file: str) -> dict:
    with open(match_cache_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    ticket_payloads = data.get("tickets", data) if isinstance(data, dict) else {}
    if not isinstance(ticket_payloads, dict):
        return {}

    normalized_cache = {}
    for ticket_key, payload in ticket_payloads.items():
        normalized_ticket_key = _normalize_ticket_key(ticket_key)
        if not normalized_ticket_key or not isinstance(payload, dict):
            continue
        normalized_cache[normalized_ticket_key] = {
            "ticket_id": payload.get("ticket_id", ""),
            "direct_site_ids": [
                _normalize_match_value(site_id)
                for site_id in payload.get("direct_site_ids", [])
                if _normalize_match_value(site_id)
            ],
            "device_ids": [
                _normalize_match_value(device_id)
                for device_id in payload.get("device_ids", [])
                if _normalize_match_value(device_id)
            ],
            "device_site_ids": [
                _normalize_match_value(site_id)
                for site_id in payload.get("device_site_ids", [])
                if _normalize_match_value(site_id)
            ],
            "matched_site_ids": [
                _normalize_match_value(site_id)
                for site_id in payload.get("matched_site_ids", [])
                if _normalize_match_value(site_id)
            ],
        }
    return normalized_cache


def _merge_unique_values(existing_values: list, new_values: list) -> list:
    merged = []
    seen = set()
    for value in list(existing_values) + list(new_values):
        normalized_value = _normalize_match_value(value)
        if not normalized_value or normalized_value in seen:
            continue
        merged.append(normalized_value)
        seen.add(normalized_value)
    return merged


def _update_ticket_match_cache(
    ticket_match_cache: dict,
    ticket_id,
    direct_site_ids: list,
    matched_device_ids: list,
    device_site_ids: list,
    matched_site_ids: list,
):
    ticket_key = _normalize_ticket_key(ticket_id)
    if not ticket_key:
        return

    entry = ticket_match_cache.setdefault(
        ticket_key,
        {
            "ticket_id": str(ticket_id).strip(),
            "direct_site_ids": [],
            "device_ids": [],
            "device_site_ids": [],
            "matched_site_ids": [],
        },
    )
    entry["direct_site_ids"] = _merge_unique_values(entry["direct_site_ids"], direct_site_ids)
    entry["device_ids"] = _merge_unique_values(entry["device_ids"], matched_device_ids)
    entry["device_site_ids"] = _merge_unique_values(entry["device_site_ids"], device_site_ids)
    entry["matched_site_ids"] = _merge_unique_values(entry["matched_site_ids"], matched_site_ids)


def _write_match_cache(match_cache_file: str, ticket_match_cache: dict, expand_sites_by_device: bool):
    os.makedirs(os.path.dirname(match_cache_file) or '.', exist_ok=True)
    payload = {
        "meta": {
            "expand_sites_by_device": bool(expand_sites_by_device),
            "ticket_count": len(ticket_match_cache),
        },
        "tickets": {
            ticket_key: {
                "ticket_id": entry.get("ticket_id", ""),
                "direct_site_ids": entry.get("direct_site_ids", []),
                "device_ids": entry.get("device_ids", []),
                "device_site_ids": entry.get("device_site_ids", []),
                "matched_site_ids": entry.get("matched_site_ids", []),
            }
            for ticket_key, entry in sorted(ticket_match_cache.items())
        },
    }
    with open(match_cache_file, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def _get_ticket_column_index(df: pd.DataFrame, ticket_field: str) -> int:
    try:
        return df.columns.get_loc(ticket_field)
    except KeyError as exc:
        raise KeyError(f"输入文件缺少 '{ticket_field}' 列") from exc


def _build_ticket_site_json(result_df, matched_sites_by_row, matched_time_entries_by_row, ticket_field: str):
    json_data = {}
    ticket_col_idx = _get_ticket_column_index(result_df, ticket_field)
    for row_idx, row in enumerate(result_df.itertuples(index=False, name=None)):
        ticket_id = row[ticket_col_idx]
        site_ids = matched_sites_by_row[row_idx]
        time_entries = matched_time_entries_by_row[row_idx] if row_idx < len(matched_time_entries_by_row) else []
        output_entry = json_data.setdefault(
            ticket_id,
            {
                "site_ids": [],
                "extracted_times": [],
                "time_details": [],
            },
        )
        output_entry["site_ids"] = _merge_exact_values(output_entry["site_ids"], site_ids)
        output_entry["extracted_times"] = _merge_exact_values(
            output_entry["extracted_times"],
            _collect_matched_times(time_entries),
        )
        _merge_time_detail_entries(output_entry["time_details"], time_entries)
    return json_data


def _normalize_header_row(header_row):
    headers = []
    used = {}
    for idx, value in enumerate(header_row):
        header = str(value).strip() if value is not None and str(value).strip() else f"Unnamed:{idx}"
        count = used.get(header, 0)
        used[header] = count + 1
        headers.append(header if count == 0 else f"{header}.{count}")
    return headers


def _merge_site_lists(primary_site_ids: list, extra_site_ids: list) -> list:
    merged = []
    seen = set()
    for site_id in list(primary_site_ids) + list(extra_site_ids):
        normalized_site_id = _normalize_match_value(site_id)
        if not normalized_site_id or normalized_site_id in seen:
            continue
        merged.append(normalized_site_id)
        seen.add(normalized_site_id)
    return merged


def _merge_exact_values(existing_values: list, new_values: list) -> list:
    merged = []
    seen = set()
    for value in list(existing_values) + list(new_values):
        exact_value = str(value).strip()
        if not exact_value or exact_value in seen:
            continue
        merged.append(exact_value)
        seen.add(exact_value)
    return merged


def _extract_ticket_date_from_key(ticket_id) -> str:
    ticket_text = str(ticket_id).strip() if ticket_id is not None else ""
    if not ticket_text:
        return ""

    matched = TICKET_DATE_PATTERN.search(ticket_text)
    if matched:
        return matched.group(1)

    matched = FALLBACK_TICKET_DATE_PATTERN.search(ticket_text)
    if matched:
        return matched.group(1)

    return ""


def _filter_times_by_ticket_date(matched_times: list, ticket_date_yyyymmdd: str) -> list:
    if not ticket_date_yyyymmdd:
        return _merge_exact_values([], matched_times)

    filtered_times = []
    for matched_time in matched_times or []:
        matched_time_text = str(matched_time).strip()
        if len(matched_time_text) < 10:
            continue
        matched_date = matched_time_text[:10].replace("-", "")
        if matched_date != ticket_date_yyyymmdd:
            continue
        filtered_times.append(matched_time_text)
    return _merge_exact_values([], filtered_times)


def _extract_time_entries_from_row(columns, row, ticket_id=None) -> list:
    time_entries = []
    seen_entries = set()
    ticket_date_yyyymmdd = _extract_ticket_date_from_key(ticket_id)

    for col_idx, value in enumerate(row):
        if value is None:
            continue

        cell_text = str(value).strip()
        if not cell_text:
            continue
        if DATETIME_FULLMATCH_PATTERN.fullmatch(cell_text):
            continue

        matched_times = DATETIME_TEXT_PATTERN.findall(cell_text)
        if not matched_times:
            continue

        deduped_times = _filter_times_by_ticket_date(matched_times, ticket_date_yyyymmdd)
        if not deduped_times:
            continue
        column_name = columns[col_idx] if col_idx < len(columns) else f"Unnamed:{col_idx}"
        entry_key = (column_name, cell_text)
        if entry_key in seen_entries:
            continue
        seen_entries.add(entry_key)
        time_entries.append(
            {
                "column": column_name,
                "cell_text": cell_text,
                "matched_times": deduped_times,
            }
        )

    return time_entries


def _collect_matched_times(time_entries: list) -> list:
    matched_times = []
    for time_entry in time_entries or []:
        if not isinstance(time_entry, dict):
            continue
        matched_times.extend(time_entry.get("matched_times", []))
    return _merge_exact_values([], matched_times)


def _merge_time_detail_entries(existing_entries: list, new_entries: list) -> None:
    existing_entry_map = {
        (str(entry.get("column", "")), str(entry.get("cell_text", ""))): entry
        for entry in existing_entries
        if isinstance(entry, dict)
    }

    for new_entry in new_entries or []:
        if not isinstance(new_entry, dict):
            continue

        column_name = str(new_entry.get("column", "")).strip()
        cell_text = str(new_entry.get("cell_text", "")).strip()
        if not column_name or not cell_text:
            continue

        entry_key = (column_name, cell_text)
        matched_times = _merge_exact_values([], new_entry.get("matched_times", []))
        if entry_key not in existing_entry_map:
            merged_entry = {
                "column": column_name,
                "cell_text": cell_text,
                "matched_times": matched_times,
            }
            existing_entries.append(merged_entry)
            existing_entry_map[entry_key] = merged_entry
            continue

        existing_entry_map[entry_key]["matched_times"] = _merge_exact_values(
            existing_entry_map[entry_key].get("matched_times", []),
            matched_times,
        )


def _attach_time_columns(result_df, matched_time_entries_by_row):
    if result_df is None:
        return None

    enriched_df = result_df.copy()
    extracted_time_lists = []
    extracted_time_details = []

    for row_idx in range(len(enriched_df)):
        time_entries = matched_time_entries_by_row[row_idx] if row_idx < len(matched_time_entries_by_row) else []
        extracted_time_lists.append("; ".join(_collect_matched_times(time_entries)))
        extracted_time_details.append(json.dumps(time_entries, ensure_ascii=False))

    enriched_df["提取时间列表"] = extracted_time_lists
    enriched_df["提取时间详情"] = extracted_time_details
    return enriched_df


def _filter_incident_tickets_rows(
    columns,
    row_iter,
    total_rows: int,
    site_device_mapping: dict,
    site_matcher: dict,
    device_site_mapping: dict = None,
    device_matcher: dict = None,
    expand_sites_by_device: bool = False,
    ticket_match_cache_input: dict = None,
    ticket_match_cache_output: dict = None,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
    progress_label: str = None,
):
    """筛选流式行数据，并仅保留命中的记录。"""
    required_device_domain, _scenario_device_label = _get_scenario_device_domain(scenario)
    filtered_indices = []
    filtered_rows = []
    matched_sites_by_row = []
    matched_time_entries_by_row = []
    stats = _empty_filter_stats()
    row_progress = ProgressBar(total_rows, progress_label or "处理工单记录", min_interval=0.05)
    ticket_col_idx = None
    try:
        ticket_col_idx = columns.index(ticket_field)
    except ValueError:
        ticket_col_idx = None

    try:
        for idx, row in enumerate(row_iter):
            stats['total'] += 1
            if len(row) < len(columns):
                row = tuple(row) + (None,) * (len(columns) - len(row))
            elif len(row) > len(columns):
                row = tuple(row[:len(columns)])

            row_text = _row_to_text(row)
            ticket_id = row[ticket_col_idx] if ticket_col_idx is not None else None
            ticket_key = _normalize_ticket_key(ticket_id)

            cached_payload = None
            if ticket_match_cache_input and ticket_key:
                cached_payload = ticket_match_cache_input.get(ticket_key)

            if cached_payload:
                direct_site_ids = list(cached_payload.get("direct_site_ids", []))
                matched_device_ids = list(cached_payload.get("device_ids", []))
                device_site_ids = list(cached_payload.get("device_site_ids", []))
                site_ids = list(cached_payload.get("matched_site_ids", []))
            else:
                direct_site_ids = extract_site_ids_from_text(row_text, site_matcher)
                matched_device_ids = []
                device_site_ids = []
                if expand_sites_by_device and device_matcher and device_site_mapping:
                    matched_device_ids = extract_outputs_from_text(row_text, device_matcher)
                    for device_id in matched_device_ids:
                        device_site_ids.extend(device_site_mapping.get(device_id, []))
                    device_site_ids = _merge_unique_values([], device_site_ids)
                site_ids = _merge_site_lists(direct_site_ids, device_site_ids)

            if ticket_match_cache_output is not None:
                _update_ticket_match_cache(
                    ticket_match_cache_output,
                    ticket_id,
                    direct_site_ids,
                    matched_device_ids,
                    device_site_ids,
                    site_ids,
                )

            if len(site_ids) < 2:
                stats['only_one_site'] += 1
                row_progress.update()
                continue

            # 检查设备类型
            valid, _devices, _required_domain_sites = check_site_devices(
                site_ids,
                site_device_mapping,
                required_device_domain,
            )

            if not valid:
                stats['missing_required_device'] += 1
                row_progress.update()
                continue

            stats['valid'] += 1
            filtered_indices.append(idx)
            filtered_rows.append(row)
            matched_sites_by_row.append(site_ids)
            matched_time_entries_by_row.append(_extract_time_entries_from_row(columns, row, ticket_id=ticket_id))
            row_progress.update()
    finally:
        row_progress.close()

    result_df = pd.DataFrame(filtered_rows, columns=columns) if filtered_rows else None
    return result_df, stats, matched_sites_by_row, matched_time_entries_by_row


def _filter_incident_tickets_df(
    df,
    site_device_mapping: dict,
    site_matcher: dict,
    device_site_mapping: dict = None,
    device_matcher: dict = None,
    expand_sites_by_device: bool = False,
    ticket_match_cache_input: dict = None,
    ticket_match_cache_output: dict = None,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
    progress_label: str = None,
):
    """筛选单个 DataFrame。"""
    row_iter = df.itertuples(index=False, name=None)
    return _filter_incident_tickets_rows(
        list(df.columns),
        row_iter,
        len(df),
        site_device_mapping,
        site_matcher,
        device_site_mapping=device_site_mapping,
        device_matcher=device_matcher,
        expand_sites_by_device=expand_sites_by_device,
        ticket_match_cache_input=ticket_match_cache_input,
        ticket_match_cache_output=ticket_match_cache_output,
        ticket_field=ticket_field,
        scenario=scenario,
        progress_label=progress_label,
    )


def _filter_incident_tickets_xlsx_stream(
    input_file: str,
    site_device_mapping: dict,
    site_matcher: dict,
    device_site_mapping: dict = None,
    device_matcher: dict = None,
    expand_sites_by_device: bool = False,
    ticket_match_cache_input: dict = None,
    ticket_match_cache_output: dict = None,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
    progress_label: str = None,
):
    """使用 openpyxl 只读模式流式读取 xlsx，避免一次性把整表读入内存。"""
    if load_workbook is None:
        raise ImportError("openpyxl 不可用，无法启用 xlsx 流式读取")

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
        )
        workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return None, _empty_filter_stats(), [], []

        columns = _normalize_header_row(header_row)
        total_rows = max((worksheet.max_row or 1) - 1, 0)
        return _filter_incident_tickets_rows(
            columns,
            row_iter,
            total_rows,
            site_device_mapping,
            site_matcher,
            device_site_mapping=device_site_mapping,
            device_matcher=device_matcher,
            expand_sites_by_device=expand_sites_by_device,
            ticket_match_cache_input=ticket_match_cache_input,
            ticket_match_cache_output=ticket_match_cache_output,
            ticket_field=ticket_field,
            scenario=scenario,
            progress_label=progress_label,
        )
    finally:
        workbook.close()


def _filter_incident_tickets_excel(
    input_file: str,
    site_device_mapping: dict,
    site_matcher: dict,
    device_site_mapping: dict = None,
    device_matcher: dict = None,
    expand_sites_by_device: bool = False,
    ticket_match_cache_input: dict = None,
    ticket_match_cache_output: dict = None,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
    progress_label: str = None,
):
    """按文件类型选择合适的 Excel 读取方式。"""
    file_ext = os.path.splitext(input_file)[1].lower()
    if file_ext in {'.xlsx', '.xlsm', '.xltx', '.xltm'} and load_workbook is not None:
        return _filter_incident_tickets_xlsx_stream(
            input_file,
            site_device_mapping,
            site_matcher,
            device_site_mapping=device_site_mapping,
            device_matcher=device_matcher,
            expand_sites_by_device=expand_sites_by_device,
            ticket_match_cache_input=ticket_match_cache_input,
            ticket_match_cache_output=ticket_match_cache_output,
            ticket_field=ticket_field,
            scenario=scenario,
            progress_label=progress_label,
        )

    df = pd.read_excel(input_file)
    print(f"原始记录数: {len(df)}")
    return _filter_incident_tickets_df(
        df,
        site_device_mapping,
        site_matcher,
        device_site_mapping=device_site_mapping,
        device_matcher=device_matcher,
        expand_sites_by_device=expand_sites_by_device,
        ticket_match_cache_input=ticket_match_cache_input,
        ticket_match_cache_output=ticket_match_cache_output,
        ticket_field=ticket_field,
        scenario=scenario,
        progress_label=progress_label,
    )


def _merge_ticket_site_json(existing_json, result_df, matched_sites_by_row, matched_time_entries_by_row, ticket_field: str):
    if result_df is None:
        return

    ticket_col_idx = _get_ticket_column_index(result_df, ticket_field)
    for row_idx, row in enumerate(result_df.itertuples(index=False, name=None)):
        ticket_id = row[ticket_col_idx]
        site_ids = matched_sites_by_row[row_idx]
        time_entries = matched_time_entries_by_row[row_idx] if row_idx < len(matched_time_entries_by_row) else []
        output_entry = existing_json.setdefault(
            ticket_id,
            {
                "site_ids": [],
                "extracted_times": [],
                "time_details": [],
            },
        )
        output_entry["site_ids"] = _merge_exact_values(output_entry["site_ids"], site_ids)
        output_entry["extracted_times"] = _merge_exact_values(
            output_entry["extracted_times"],
            _collect_matched_times(time_entries),
        )
        _merge_time_detail_entries(output_entry["time_details"], time_entries)


def _count_total_sites_from_ticket_json(ticket_json: dict) -> int:
    total = 0
    for payload in (ticket_json or {}).values():
        if isinstance(payload, dict):
            total += len(payload.get("site_ids", []))
    return total


def _filter_incident_tickets_file(
    input_file: str,
    site_device_mapping: dict,
    site_matcher: dict,
    output_file: str,
    json_output_file: str = None,
    match_cache_output_file: str = None,
    ticket_match_cache_input: dict = None,
    ticket_match_cache_output: dict = None,
    device_site_mapping: dict = None,
    device_matcher: dict = None,
    expand_sites_by_device: bool = False,
    json_only: bool = False,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
):
    """筛选单个 Excel 文件。"""
    _print_file_start(input_file)
    _required_device_domain, scenario_device_label = _get_scenario_device_domain(scenario)

    result_df, stats, matched_sites_by_row, matched_time_entries_by_row = _filter_incident_tickets_excel(
        input_file,
        site_device_mapping,
        site_matcher,
        device_site_mapping=device_site_mapping,
        device_matcher=device_matcher,
        expand_sites_by_device=expand_sites_by_device,
        ticket_match_cache_input=ticket_match_cache_input,
        ticket_match_cache_output=ticket_match_cache_output,
        ticket_field=ticket_field,
        scenario=scenario,
        progress_label=f"处理记录 {os.path.basename(input_file)}",
    )

    _print_stats("筛选统计", stats, scenario_device_label)

    # 输出结果
    if result_df is not None:
        result_df = _attach_time_columns(result_df, matched_time_entries_by_row)
        if not json_only:
            os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
            result_df.to_excel(output_file, index=False)

        if json_output_file:
            os.makedirs(os.path.dirname(json_output_file) or '.', exist_ok=True)
            json_data = _build_ticket_site_json(
                result_df,
                matched_sites_by_row,
                matched_time_entries_by_row,
                ticket_field,
            )
            with open(json_output_file, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
        _print_section("处理结果")
        _print_output_result(len(result_df), None if json_only else output_file, json_output_file)
        _print_key_values([("输出记录关联站点总数", _count_total_sites(matched_sites_by_row))])
        if match_cache_output_file and ticket_match_cache_output is not None:
            _write_match_cache(match_cache_output_file, ticket_match_cache_output, expand_sites_by_device)
            _print_key_values([("匹配中间结果", match_cache_output_file)])

        return result_df, stats
    else:
        if match_cache_output_file and ticket_match_cache_output is not None:
            _write_match_cache(match_cache_output_file, ticket_match_cache_output, expand_sites_by_device)
        _print_section("处理结果")
        if match_cache_output_file and ticket_match_cache_output is not None:
            _print_key_values([("匹配中间结果", match_cache_output_file)])
        print("没有满足条件的记录")
        return None, stats


def filter_incident_tickets(
    input_file: str,
    site_device_file: str,
    output_file: str,
    json_output_file: str = None,
    match_cache_input_file: str = None,
    match_cache_output_file: str = None,
    ne_graph_file: str = None,
    expand_sites_by_device: bool = False,
    json_only: bool = False,
    ticket_field: str = "工单ID",
    scenario: str = "transmission",
):
    """筛选满足条件的Incident Ticket记录"""
    _required_device_domain, scenario_device_label = _get_scenario_device_domain(scenario)
    # 加载站点设备映射
    site_device_mapping = load_site_device_mapping(site_device_file)
    known_site_ids = build_known_site_ids(site_device_mapping)
    site_matcher = build_site_matcher(known_site_ids)
    device_matcher = None
    device_site_mapping = {}
    if expand_sites_by_device:
        if not ne_graph_file:
            raise ValueError("开启 expand-sites-by-device 时，必须提供 --ne-graph")
        device_site_mapping = load_device_site_mapping(ne_graph_file)
        device_matcher = build_device_matcher(device_site_mapping)
    ticket_match_cache_input = _load_match_cache(match_cache_input_file) if match_cache_input_file else {}
    ticket_match_cache_output = dict(ticket_match_cache_input)
    if not match_cache_output_file:
        match_cache_output_file = _derive_match_cache_output_path(output_file)
    _print_section("初始化")
    init_items = [
        ("筛选场景", scenario),
        ("要求设备域", scenario_device_label),
        ("已加载站点数", len(site_device_mapping)),
    ]
    if expand_sites_by_device:
        init_items.extend([
            ("设备补站点", "开启"),
            ("设备关键词数", len(device_site_mapping)),
        ])
    if match_cache_input_file:
        init_items.append(("匹配中间结果输入", match_cache_input_file))
    if match_cache_output_file:
        init_items.append(("匹配中间结果输出", match_cache_output_file))
    init_items.append(("工单字段", ticket_field))
    _print_key_values(init_items)
    return _filter_incident_tickets_file(
        input_file,
        site_device_mapping,
        site_matcher,
        output_file,
        json_output_file,
        match_cache_output_file=match_cache_output_file,
        ticket_match_cache_input=ticket_match_cache_input,
        ticket_match_cache_output=ticket_match_cache_output,
        device_site_mapping=device_site_mapping,
        device_matcher=device_matcher,
        expand_sites_by_device=expand_sites_by_device,
        json_only=json_only,
        ticket_field=ticket_field,
        scenario=scenario,
    )


def _iter_incident_input_files(input_path: str):
    if os.path.isdir(input_path):
        for root, dirnames, filenames in os.walk(input_path):
            dirnames.sort()
            for filename in sorted(filenames):
                if filename.lower().endswith(('.xlsx', '.xls')):
                    yield os.path.join(root, filename)
        return

    yield input_path

def main():
    parser = argparse.ArgumentParser(description='筛选Incident Ticket记录')
    parser.add_argument(
        '-i', '--input',
        default=DEFAULT_INCIDENT_TICKET_XLSX,
        help=f'输入的Excel文件，默认: {ticket_resource_display("Incident Ticket_20260201-20260318.xlsx")}'
    )
    parser.add_argument(
        '-s', '--site-device',
        default=SITE_DEVICE_COUNTS_JSON,
        help=f'站点设备映射JSON文件，默认: {resource_display("site_device_counts.json")}'
    )
    parser.add_argument(
        '-o', '--output',
        default='filtered_incident_tickets.xlsx',
        help='输出的Excel文件'
    )
    parser.add_argument(
        '-j', '--json-output',
        help='JSON输出文件（可选），格式：{工单字段值: {site_ids, extracted_times, time_details}}'
    )
    parser.add_argument(
        '--ticket-field',
        default='工单ID',
        help='输入 Excel 中的工单字段列名，默认: 工单ID'
    )
    parser.add_argument(
        '--scenario',
        choices=sorted(SCENARIO_DEVICE_DOMAINS),
        default='transmission',
        help='筛选场景：transmission 要求至少两个匹配站点包含 Transmission 设备；data 要求至少两个匹配站点包含 Data 设备'
    )
    parser.add_argument(
        '--ne-graph',
        default=NE_GRAPH_JSON,
        help=f'ne_graph.json 文件；默认: {resource_display("ne_graph.json")}；开启设备补站点时需要可用'
    )
    parser.add_argument(
        '--expand-sites-by-device',
        action='store_true',
        help='在直接匹配站点ID之外，再通过匹配上的设备ID关联出站点'
    )
    parser.add_argument(
        '--match-cache-input',
        help='已存在的工单匹配中间结果 JSON；提供后优先直接使用其中的工单匹配站点/设备'
    )
    parser.add_argument(
        '--match-cache-output',
        help='输出工单匹配中间结果 JSON；默认随 Excel 输出生成同名 .match_cache.json'
    )
    parser.add_argument(
        '--json-only',
        action='store_true',
        help='只输出 JSON 和匹配中间结果，不生成 Excel'
    )

    args = parser.parse_args()
    _required_device_domain, scenario_device_label = _get_scenario_device_domain(args.scenario)

    site_device_mapping = load_site_device_mapping(args.site_device)
    known_site_ids = build_known_site_ids(site_device_mapping)
    site_matcher = build_site_matcher(known_site_ids)
    device_matcher = None
    device_site_mapping = {}
    if args.expand_sites_by_device:
        if not args.ne_graph:
            raise ValueError("开启 expand-sites-by-device 时，必须提供 --ne-graph")
        device_site_mapping = load_device_site_mapping(args.ne_graph)
        device_matcher = build_device_matcher(device_site_mapping)
    ticket_match_cache_input = _load_match_cache(args.match_cache_input) if args.match_cache_input else {}
    match_cache_output_file = args.match_cache_output or _derive_match_cache_output_path(args.output)
    _print_section("初始化")
    init_items = [
        ("筛选场景", args.scenario),
        ("要求设备域", scenario_device_label),
        ("已加载站点数", len(site_device_mapping)),
    ]
    if args.expand_sites_by_device:
        init_items.extend([
            ("设备补站点", "开启"),
            ("设备关键词数", len(device_site_mapping)),
        ])
    if args.json_only:
        init_items.append(("仅输出 JSON", "开启"))
    if args.match_cache_input:
        init_items.append(("匹配中间结果输入", args.match_cache_input))
    init_items.append(("匹配中间结果输出", match_cache_output_file))
    init_items.append(("工单字段", args.ticket_field))
    _print_key_values(init_items)

    input_files = list(_iter_incident_input_files(args.input))
    if not input_files:
        print("没有找到可处理的 Excel 文件")
        return

    if os.path.isdir(args.input):
        aggregate_stats = _empty_filter_stats()
        processed_files = 0
        aggregated_result_dfs = []
        aggregated_json = {}
        aggregated_match_cache = dict(ticket_match_cache_input)
        file_progress = ProgressBar(len(input_files), "处理输入文件", min_interval=0.05)

        try:
            for input_file in input_files:
                _print_file_start(input_file)

                result_df, stats, matched_sites_by_row, matched_time_entries_by_row = _filter_incident_tickets_excel(
                    input_file,
                    site_device_mapping,
                    site_matcher,
                    device_site_mapping=device_site_mapping,
                    device_matcher=device_matcher,
                    expand_sites_by_device=args.expand_sites_by_device,
                    ticket_match_cache_input=aggregated_match_cache,
                    ticket_match_cache_output=aggregated_match_cache,
                    ticket_field=args.ticket_field,
                    scenario=args.scenario,
                    progress_label=f"处理记录 {os.path.basename(input_file)}",
                )

                _print_stats("筛选统计", stats, scenario_device_label)

                processed_files += 1
                for key in aggregate_stats:
                    aggregate_stats[key] += stats.get(key, 0)
                if result_df is not None:
                    result_df = _attach_time_columns(result_df, matched_time_entries_by_row)
                    aggregated_result_dfs.append(result_df)
                    _merge_ticket_site_json(
                        aggregated_json,
                        result_df,
                        matched_sites_by_row,
                        matched_time_entries_by_row,
                        args.ticket_field,
                    )
                file_progress.update()
        finally:
            file_progress.close()

        _print_section("目录处理汇总")
        _print_key_values([
            ("处理文件数", processed_files),
            ("总记录数", aggregate_stats['total']),
            ("命中记录数", aggregate_stats['valid']),
            ("不足 2 个站点", aggregate_stats['only_one_site']),
            (f"具备 {scenario_device_label} 的站点不足 2 个", aggregate_stats['missing_required_device']),
            ("输出记录关联站点总数", _count_total_sites_from_ticket_json(aggregated_json)),
        ])

        if aggregated_result_dfs:
            final_result_df = pd.concat(aggregated_result_dfs, ignore_index=True)
            _print_section("汇总结果")
            if not args.json_only:
                os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
                final_result_df.to_excel(args.output, index=False)
            _print_output_result(len(final_result_df), None if args.json_only else args.output, args.json_output, aggregated=True)
        else:
            _print_section("汇总结果")
            print("没有满足条件的记录")

        if args.json_output:
            os.makedirs(os.path.dirname(args.json_output) or '.', exist_ok=True)
            with open(args.json_output, 'w', encoding='utf-8') as f:
                json.dump(aggregated_json, f, ensure_ascii=False, indent=2)
        _write_match_cache(match_cache_output_file, aggregated_match_cache, args.expand_sites_by_device)
        _print_key_values([("匹配中间结果", match_cache_output_file)])
        return

    filter_incident_tickets(
        args.input,
        args.site_device,
        args.output,
        args.json_output,
        match_cache_input_file=args.match_cache_input,
        match_cache_output_file=match_cache_output_file,
        ne_graph_file=args.ne_graph,
        expand_sites_by_device=args.expand_sites_by_device,
        json_only=args.json_only,
        ticket_field=args.ticket_field,
        scenario=args.scenario,
    )


if __name__ == '__main__':
    main()
