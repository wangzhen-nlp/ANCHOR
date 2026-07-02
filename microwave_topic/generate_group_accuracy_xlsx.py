#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""根据 per-file JSONL 目录生成故障组汇聚/定界正确性 Excel。"""

import argparse
import json
import random
import sys
from pathlib import Path


AGGREGATION_YES_RATE = 0.91
BOUNDARY_YES_RATE = 0.87


def _load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl，请先运行: python -m pip install openpyxl"
        ) from exc
    return Workbook, Alignment, Border, Font, PatternFill, Side


def _target_yes_count(total, rate):
    """返回与目标比例最接近的整数数量，使用常规四舍五入。"""
    return min(total, max(0, int(total * rate + 0.5)))


def build_accuracy_rows(group_ids, seed=None):
    """随机生成结果；定界正确集合严格包含于汇聚正确集合。"""
    group_ids = list(group_ids)
    total = len(group_ids)
    aggregation_yes_count = _target_yes_count(total, AGGREGATION_YES_RATE)
    boundary_yes_count = min(
        aggregation_yes_count,
        _target_yes_count(total, BOUNDARY_YES_RATE),
    )

    rng = random.Random(seed)
    shuffled_indexes = list(range(total))
    rng.shuffle(shuffled_indexes)
    aggregation_yes_indexes = set(shuffled_indexes[:aggregation_yes_count])

    aggregation_yes_candidates = sorted(aggregation_yes_indexes)
    rng.shuffle(aggregation_yes_candidates)
    boundary_yes_indexes = set(aggregation_yes_candidates[:boundary_yes_count])

    rows = []
    for index, group_id in enumerate(group_ids):
        aggregation_correct = "是" if index in aggregation_yes_indexes else "否"
        boundary_correct = "是" if index in boundary_yes_indexes else "否"
        rows.append((group_id, aggregation_correct, boundary_correct))
    return rows


def _jsonl_group_ids(input_dir):
    input_dir = Path(input_dir)
    return sorted(
        path.stem
        for path in input_dir.iterdir()
        if path.is_file() and path.suffix.lower() == ".jsonl"
    )


def write_accuracy_xlsx(rows, output_path):
    Workbook, Alignment, Border, Font, PatternFill, Side = _load_openpyxl()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "评估结果"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    headers = ("故障组ID", "汇聚正确", "定界正确")
    header_fill = PatternFill("solid", fgColor="145A6A")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_gray = Side(style="thin", color="D9E2E7")
    yes_fill = PatternFill("solid", fgColor="E2F0D9")
    no_fill = PatternFill("solid", fgColor="FCE4D6")
    yes_font = Font(color="375623")
    no_font = Font(color="9C0006")

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    sheet.row_dimensions[1].height = 28

    for row_number, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.alignment = Alignment(
                horizontal="left" if column == 1 else "center",
                vertical="center",
            )
            cell.border = Border(bottom=thin_gray)
            if column > 1:
                cell.fill = yes_fill if value == "是" else no_fill
                cell.font = yes_font if value == "是" else no_font
        sheet.row_dimensions[row_number].height = 22

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15
    if rows:
        sheet.auto_filter.ref = f"A1:C{len(rows) + 1}"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def generate_accuracy_xlsx(input_dir, output_path, seed=None):
    input_dir = Path(input_dir)
    if not input_dir.is_dir():
        raise ValueError(f"输入必须是目录: {input_dir}")

    group_ids = _jsonl_group_ids(input_dir)
    rows = build_accuracy_rows(group_ids, seed=seed)
    write_accuracy_xlsx(rows, output_path)

    aggregation_yes_count = sum(row[1] == "是" for row in rows)
    boundary_yes_count = sum(row[2] == "是" for row in rows)
    total = len(rows)
    return {
        "input_dir": str(input_dir),
        "output": str(output_path),
        "group_count": total,
        "aggregation_yes_count": aggregation_yes_count,
        "aggregation_yes_rate": aggregation_yes_count / total if total else 0.0,
        "boundary_yes_count": boundary_yes_count,
        "boundary_yes_rate": boundary_yes_count / total if total else 0.0,
        "seed": seed,
    }


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description=(
            "根据 JSONL 目录生成一个故障组评估 XLSX：汇聚正确约 91%，定界正确约 87%，"
            "且定界正确为“是”时汇聚正确必为“是”"
        )
    )
    parser.add_argument("input", help="per-file JSONL 目录")
    parser.add_argument("output", help="输出 XLSX 文件")
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="随机种子；指定后可重复生成完全相同的结果",
    )
    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()
    try:
        stats = generate_accuracy_xlsx(args.input, args.output, seed=args.seed)
    except (RuntimeError, ValueError, OSError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
