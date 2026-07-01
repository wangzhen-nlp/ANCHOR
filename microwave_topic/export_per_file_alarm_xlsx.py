#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 complete_group_topology.py --per-file 的 JSONL 逐个导出为层级告警 Excel。"""

import argparse
import json
import re
import sys
from pathlib import Path


PRIMARY_COLUMNS = (
    "站点ID",
    "设备厂家",
    "告警源",
    "告警标题",
    "告警首次发生时间",
    "告警清除时间",
    "告警最后发生时间",
)

OTHER_MAIN_COLUMNS = (
    "故障组ID",
    "告警序号",
    "告警编码ID",
    "告警标准名",
    "设备名称",
    "设备域",
    "站点名称",
    "站点类型",
    "区域ID",
    "告警级别",
    "告警状态",
    "工单号",
    "经度",
    "纬度",
)

MAIN_COLUMNS = PRIMARY_COLUMNS + OTHER_MAIN_COLUMNS

FIELD_ALIASES = {
    "告警编码ID": ("告警编码ID", "告警ID", "eid", "alarm_id", "event_id", "id"),
    "告警标题": ("告警标题", "alarm", "alarm_type", "alarm_title", "title"),
    "告警标准名": ("告警标准名", "告警标准化名称", "standard_alarm_name", "standard_name"),
    "告警源": ("告警源", "alarm_source", "ne_id", "source"),
    "设备名称": ("设备名称", "网元名称", "device_name", "ne_name"),
    "设备域": ("设备域", "domain", "网络专业", "告警源专业", "专业"),
    "站点ID": ("站点ID", "site_id", "node", "site"),
    "站点名称": ("站点名称", "site_name"),
    "站点类型": ("站点类型", "site_type"),
    "区域ID": ("区域ID", "region_id"),
    "告警级别": ("告警级别", "告警等级", "级别", "severity"),
    "告警状态": ("告警状态", "状态", "alarm_status", "status"),
    "告警首次发生时间": (
        "告警首次发生时间",
        "告警发生时间",
        "首次发生时间",
        "发生时间",
        "alarm_time",
        "time",
        "ts",
    ),
    "告警清除时间": ("告警清除时间", "清除时间", "alarm_clear_time", "clear_time"),
    "告警最后发生时间": ("告警最后发生时间", "最后发生时间", "last_occurrence_time"),
    "工单号": ("工单号", "ticket_id", "work_order_id"),
    "设备厂家": ("设备厂家", "设备厂家名称", "厂家", "manufacturer"),
    "经度": ("经度", "longitude", "lon", "lng"),
    "纬度": ("纬度", "latitude", "lat"),
}

CONSUMED_ALARM_FIELDS = {field for aliases in FIELD_ALIASES.values() for field in aliases}


def _load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "缺少依赖 openpyxl，请先运行: python -m pip install openpyxl"
        ) from exc
    return Workbook, Alignment, Border, Font, PatternFill, Side


def _normalize_text(value):
    return "" if value is None else str(value).strip()


def _first_value(record, fields):
    if not isinstance(record, dict):
        return ""
    for field in fields:
        value = record.get(field)
        if value not in (None, ""):
            return value
    return ""


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    if isinstance(value, str) and value.startswith("="):
        return "'" + value
    return value


def _group_id(group):
    match_info = group.get("match_info") if isinstance(group.get("match_info"), dict) else {}
    return _normalize_text(
        group.get("uuid") or group.get("故障组ID") or match_info.get("uuid") or ""
    )


def _alarm_records(group):
    """只选一个最可信来源，避免 alarms/symptoms/ne_info 的可视化副本重复。"""
    alarms = [item for item in group.get("alarms") or [] if isinstance(item, dict)]
    if alarms:
        return [(item, "") for item in alarms]

    symptoms = [item for item in group.get("symptoms") or [] if isinstance(item, dict)]
    if symptoms:
        return [(item, "") for item in symptoms]

    match_info = group.get("match_info") if isinstance(group.get("match_info"), dict) else {}
    symptoms = [item for item in match_info.get("symptoms") or [] if isinstance(item, dict)]
    if symptoms:
        return [(item, "") for item in symptoms]

    result = []
    ne_info = group.get("ne_info") if isinstance(group.get("ne_info"), dict) else {}
    for ne_id, info in ne_info.items():
        if not isinstance(info, dict):
            continue
        for alarm in info.get("alarm") or []:
            if isinstance(alarm, dict):
                result.append((alarm, _normalize_text(ne_id)))
    return result


def _device_context(group, alarm, fallback_ne_id):
    ne_id = _normalize_text(_first_value(alarm, FIELD_ALIASES["告警源"])) or fallback_ne_id
    ne_info = group.get("ne_info") if isinstance(group.get("ne_info"), dict) else {}
    info = ne_info.get(ne_id, {}) if ne_id else {}
    return ne_id, info if isinstance(info, dict) else {}


def group_alarm_rows(group):
    rows = []
    group_uuid = _group_id(group)
    for index, (alarm, fallback_ne_id) in enumerate(_alarm_records(group), start=1):
        ne_id, info = _device_context(group, alarm, fallback_ne_id)
        row = {column: "" for column in MAIN_COLUMNS}
        row["故障组ID"] = group_uuid
        row["告警序号"] = index
        for column, aliases in FIELD_ALIASES.items():
            row[column] = _excel_value(_first_value(alarm, aliases))

        row["告警源"] = row["告警源"] or ne_id
        row["设备名称"] = row["设备名称"] or _excel_value(info.get("name", ""))
        row["设备域"] = row["设备域"] or _excel_value(info.get("domain", ""))
        row["站点ID"] = row["站点ID"] or _excel_value(info.get("site_id", ""))
        row["站点名称"] = row["站点名称"] or _excel_value(info.get("site_name", ""))
        row["站点类型"] = row["站点类型"] or _excel_value(info.get("site_type", ""))
        row["区域ID"] = row["区域ID"] or _excel_value(info.get("region_id", ""))
        row["设备厂家"] = row["设备厂家"] or _excel_value(info.get("manufacturer", ""))
        row["经度"] = row["经度"] or _excel_value(info.get("longitude", ""))
        row["纬度"] = row["纬度"] or _excel_value(info.get("latitude", ""))

        for raw_key, raw_value in alarm.items():
            key = _normalize_text(raw_key)
            if key and key not in CONSUMED_ALARM_FIELDS and key not in row:
                row[key] = _excel_value(raw_value)
        row["__original_order"] = index
        rows.append(row)
    return rows


def _natural_key(value):
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", _normalize_text(value))
    )


def _sorted_rows(rows):
    return sorted(
        rows,
        key=lambda row: (
            _natural_key(row.get("站点ID")),
            _natural_key(row.get("设备厂家")),
            _natural_key(row.get("告警源")),
            row.get("__original_order", 0),
        ),
    )


def _hierarchy_spans(rows, fields):
    for level, field in enumerate(fields):
        start = 0
        while start < len(rows):
            end = start
            while end + 1 < len(rows) and all(
                _normalize_text(rows[end + 1].get(parent_field))
                == _normalize_text(rows[start].get(parent_field))
                for parent_field in fields[: level + 1]
            ):
                end += 1
            yield level, field, start, end
            start = end + 1


def _read_groups(input_path):
    groups = []
    with Path(input_path).open("r", encoding="utf-8-sig") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            line = raw_line.strip()
            if not line:
                continue
            try:
                group = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{input_path} 第 {line_number} 行 JSON 解析失败: {exc}") from exc
            if not isinstance(group, dict):
                raise ValueError(f"{input_path} 第 {line_number} 行必须是 JSON 对象")
            groups.append(group)
    return groups


def build_alarm_workbook(input_path):
    Workbook, Alignment, Border, Font, PatternFill, Side = _load_openpyxl()
    rows = _sorted_rows(
        [row for group in _read_groups(input_path) for row in group_alarm_rows(group)]
    )
    extra_columns = []
    for row in rows:
        for key in row:
            if key != "__original_order" and key not in MAIN_COLUMNS and key not in extra_columns:
                extra_columns.append(key)
    columns = list(MAIN_COLUMNS) + extra_columns

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "告警明细"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "D2"

    header_fill = PatternFill("solid", fgColor="145A6A")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="AABBC3")
    light_side = Side(style="thin", color="D9E2E7")

    for column_index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(top=thin_side, bottom=thin_side)
    sheet.row_dimensions[1].height = 30

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row.get(column, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=light_side)
        sheet.row_dimensions[row_index].height = 24

    hierarchy_colors = ("E8F2F5", "EEF5E8", "FFF5DA")
    for level, _field, start, end in _hierarchy_spans(rows, PRIMARY_COLUMNS[:3]):
        excel_start = start + 2
        excel_end = end + 2
        column_index = level + 1
        for row_index in range(excel_start, excel_end + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.fill = PatternFill("solid", fgColor=hierarchy_colors[level])
            cell.font = Font(bold=level < 2, color="24363D")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        if excel_end > excel_start:
            sheet.merge_cells(
                start_row=excel_start,
                start_column=column_index,
                end_row=excel_end,
                end_column=column_index,
            )

    widths = (18, 15, 24, 38, 21, 21, 21)
    from openpyxl.utils import get_column_letter

    for index in range(1, len(columns) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = (
            widths[index - 1] if index <= len(widths) else 17
        )
    return workbook, len(rows)


def export_jsonl_file(input_path, output_path):
    workbook, alarm_count = build_alarm_workbook(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return alarm_count


def export_per_file_directory(input_dir, output_dir=None):
    input_dir = Path(input_dir)
    output_dir = Path(output_dir) if output_dir else input_dir
    input_files = sorted(input_dir.glob("*.jsonl"))
    stats = {
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "input_file_count": len(input_files),
        "output_file_count": 0,
        "alarm_count": 0,
    }
    for input_path in input_files:
        stats["alarm_count"] += export_jsonl_file(
            input_path, output_dir / f"{input_path.stem}.xlsx"
        )
        stats["output_file_count"] += 1
    return stats


def build_arg_parser():
    parser = argparse.ArgumentParser(
        description="把 complete_group_topology.py --per-file 输出的每个 JSONL 导出为层级 Excel"
    )
    parser.add_argument("input", help="per-file JSONL 目录，也支持单个 JSONL 文件")
    parser.add_argument(
        "output",
        nargs="?",
        help="XLSX 输出目录或单个 XLSX 文件；省略时输出到 JSONL 所在目录",
    )
    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()
    input_path = Path(args.input)
    if not input_path.exists():
        parser.error(f"输入不存在: {input_path}")

    try:
        if input_path.is_file():
            output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")
            alarm_count = export_jsonl_file(input_path, output_path)
            stats = {
                "input": str(input_path),
                "output": str(output_path),
                "output_file_count": 1,
                "alarm_count": alarm_count,
            }
        else:
            stats = export_per_file_directory(input_path, args.output)
    except (RuntimeError, ValueError, OSError) as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
