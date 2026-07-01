import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from microwave_topic.export_per_file_alarm_xlsx import (
    export_jsonl_file,
    export_per_file_directory,
)


class ExportPerFileAlarmXlsxTest(unittest.TestCase):
    def setUp(self):
        self.group = {
            "uuid": "GROUP-1",
            "alarms": [
                {"告警编码ID": "A-2", "告警标题": "链路中断", "告警源": "NE-2"},
                {"告警编码ID": "A-1", "告警标题": "网元断连", "告警源": "NE-1"},
                {"告警编码ID": "A-4", "告警标题": "设备离线", "告警源": "NE-1"},
                {"告警编码ID": "A-3", "告警标题": "电源告警", "告警源": "NE-3"},
            ],
            "symptoms": [{"eid": "A-1", "alarm": "网元断连", "alarm_source": "NE-1"}],
            "ne_info": {
                "NE-1": {"site_id": "SITE-1", "manufacturer": "HUAWEI"},
                "NE-2": {"site_id": "SITE-1", "manufacturer": "HUAWEI"},
                "NE-3": {"site_id": "SITE-2", "manufacturer": "ZTE"},
            },
        }

    def test_exports_sorted_rows_and_merges_each_hierarchy_level(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "GROUP-1.jsonl"
            output_path = Path(tmpdir) / "GROUP-1.xlsx"
            input_path.write_text(json.dumps(self.group, ensure_ascii=False) + "\n", encoding="utf-8")
            self.assertEqual(export_jsonl_file(input_path, output_path), 4)
            workbook = load_workbook(output_path)
            sheet = workbook["告警明细"]

        self.assertEqual(
            [sheet.cell(1, column).value for column in range(1, 8)],
            [
                "站点ID",
                "设备厂家",
                "告警源",
                "告警标题",
                "告警首次发生时间",
                "告警清除时间",
                "告警最后发生时间",
            ],
        )
        self.assertEqual(sheet["A2"].value, "SITE-1")
        self.assertEqual(sheet["C2"].value, "NE-1")
        self.assertEqual(sheet["C4"].value, "NE-2")
        self.assertEqual(sheet["A5"].value, "SITE-2")
        self.assertIn("A2:A4", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("B2:B4", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("C2:C3", {str(item) for item in sheet.merged_cells.ranges})

    def test_writes_one_same_name_xlsx_per_jsonl(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "jsonl"
            output_dir = Path(tmpdir) / "xlsx"
            input_dir.mkdir()
            (input_dir / "GROUP-1.jsonl").write_text(
                json.dumps(self.group, ensure_ascii=False) + "\n", encoding="utf-8"
            )
            stats = export_per_file_directory(input_dir, output_dir)
            self.assertTrue((output_dir / "GROUP-1.xlsx").exists())

        self.assertEqual(stats["input_file_count"], 1)
        self.assertEqual(stats["output_file_count"], 1)
        self.assertEqual(stats["alarm_count"], 4)


if __name__ == "__main__":
    unittest.main()
