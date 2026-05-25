#!/usr/bin/env python3
"""Count data rows in CSV/JSONL/Excel files under a directory, including ZIP members.

By default one header row is subtracted from every non-empty CSV file and every
non-empty Excel sheet. JSONL/NDJSON files are counted as headerless records.
"""

from __future__ import annotations

from argparse import ArgumentParser
from dataclasses import dataclass
from io import BytesIO, TextIOWrapper
import csv
import json
import os
import posixpath
from pathlib import Path, PurePosixPath
import re
from xml.etree import ElementTree as ET
import zipfile

if __package__ in (None, ""):
    from _script_env import ensure_repo_root

    ensure_repo_root(1)

from alarm_tools.progress_utils import ProgressBar


CSV_EXTS = {".csv"}
JSONL_EXTS = {".jsonl", ".ndjson"}
EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
ZIP_EXTS = {".zip"}
DEFAULT_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030")
PROGRESS_PATH_MAX_LEN = 90


try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on local environment
    load_workbook = None


@dataclass
class CountRecord:
    path: str
    kind: str
    physical_rows: int
    data_rows: int
    header_rows: int
    sheet: str = ""
    error: str = ""

    def to_dict(self):
        return {
            "path": self.path,
            "kind": self.kind,
            "sheet": self.sheet,
            "physical_rows": self.physical_rows,
            "header_rows": self.header_rows,
            "data_rows": self.data_rows,
            "error": self.error,
        }


def _suffix(path):
    return Path(str(path)).suffix.lower()


def _is_supported_table(path):
    return _suffix(path) in CSV_EXTS | JSONL_EXTS | EXCEL_EXTS


def _is_zip(path):
    return _suffix(path) in ZIP_EXTS


def _is_nonempty_row(row):
    return any(str(cell).strip() for cell in row if cell is not None)


def _data_rows(physical_rows, header_rows):
    return max(int(physical_rows) - int(header_rows), 0) if physical_rows > 0 else 0


def _count_csv_stream(text_stream, *, header_rows):
    reader = csv.reader(text_stream)
    physical_rows = 0
    for row in reader:
        if _is_nonempty_row(row):
            physical_rows += 1
    return physical_rows, _data_rows(physical_rows, header_rows)


def _count_csv_file(path, *, header_rows, encodings=DEFAULT_ENCODINGS):
    last_error = None
    for encoding in encodings:
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                physical_rows, data_rows = _count_csv_stream(handle, header_rows=header_rows)
            return CountRecord(
                path=str(path),
                kind="csv",
                physical_rows=physical_rows,
                data_rows=data_rows,
                header_rows=min(header_rows, physical_rows),
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"cannot decode with encodings: {', '.join(encodings)}; last={last_error}",
    )


def _count_csv_bytes(data, logical_path, *, header_rows, encodings=DEFAULT_ENCODINGS):
    last_error = None
    for encoding in encodings:
        try:
            with TextIOWrapper(BytesIO(data), encoding=encoding, newline="") as handle:
                physical_rows, data_rows = _count_csv_stream(handle, header_rows=header_rows)
            return CountRecord(
                path=logical_path,
                kind="csv",
                physical_rows=physical_rows,
                data_rows=data_rows,
                header_rows=min(header_rows, physical_rows),
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"cannot decode with encodings: {', '.join(encodings)}; last={last_error}",
    )


def _count_jsonl_stream(text_stream):
    physical_rows = 0
    for line in text_stream:
        if line.strip():
            physical_rows += 1
    return physical_rows


def _count_jsonl_file(path, *, encodings=DEFAULT_ENCODINGS):
    last_error = None
    for encoding in encodings:
        try:
            with open(path, "r", encoding=encoding) as handle:
                physical_rows = _count_jsonl_stream(handle)
            return CountRecord(
                path=str(path),
                kind="jsonl",
                physical_rows=physical_rows,
                data_rows=physical_rows,
                header_rows=0,
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise UnicodeDecodeError(
        "jsonl",
        b"",
        0,
        1,
        f"cannot decode with encodings: {', '.join(encodings)}; last={last_error}",
    )


def _count_jsonl_bytes(data, logical_path, *, encodings=DEFAULT_ENCODINGS):
    last_error = None
    for encoding in encodings:
        try:
            with TextIOWrapper(BytesIO(data), encoding=encoding) as handle:
                physical_rows = _count_jsonl_stream(handle)
            return CountRecord(
                path=logical_path,
                kind="jsonl",
                physical_rows=physical_rows,
                data_rows=physical_rows,
                header_rows=0,
            )
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise UnicodeDecodeError(
        "jsonl",
        b"",
        0,
        1,
        f"cannot decode with encodings: {', '.join(encodings)}; last={last_error}",
    )


def _iter_worksheet_rows(worksheet):
    for row in worksheet.iter_rows(values_only=True):
        if _is_nonempty_row(row):
            yield row


def _count_workbook(workbook, logical_path, *, header_rows, first_sheet=False):
    records = []
    worksheets = workbook.worksheets[:1] if first_sheet else workbook.worksheets
    for worksheet in worksheets:
        physical_rows = sum(1 for _ in _iter_worksheet_rows(worksheet))
        records.append(
            CountRecord(
                path=logical_path,
                kind="excel",
                sheet=worksheet.title,
                physical_rows=physical_rows,
                data_rows=_data_rows(physical_rows, header_rows),
                header_rows=min(header_rows, physical_rows),
            )
        )
    return records


def _count_xlsx_file(path, *, header_rows, first_sheet=False):
    if load_workbook is None:
        return _count_xlsx_zip(path, str(path), header_rows=header_rows, first_sheet=first_sheet)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _count_workbook(
            workbook,
            str(path),
            header_rows=header_rows,
            first_sheet=first_sheet,
        )
    finally:
        workbook.close()


def _count_xlsx_bytes(data, logical_path, *, header_rows, first_sheet=False):
    if load_workbook is None:
        return _count_xlsx_zip(BytesIO(data), logical_path, header_rows=header_rows, first_sheet=first_sheet)
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        return _count_workbook(
            workbook,
            logical_path,
            header_rows=header_rows,
            first_sheet=first_sheet,
        )
    finally:
        workbook.close()


def _xml_local_name(tag):
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _xlsx_sheet_number(member_name):
    match = re.search(r"sheet(\d+)\.xml$", member_name)
    return int(match.group(1)) if match else 10**9


def _xlsx_sheet_names(zf):
    names_by_target = {}
    try:
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {}
        for rel in rels_root:
            rel_id = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if not rel_id or not target:
                continue
            rel_targets[rel_id] = posixpath.normpath(posixpath.join("xl", target))

        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        for elem in workbook_root.iter():
            if _xml_local_name(elem.tag) != "sheet":
                continue
            rel_id = elem.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            target = rel_targets.get(rel_id)
            if target:
                names_by_target[target] = elem.attrib.get("name", "") or PurePosixPath(target).stem
    except Exception:
        return {}
    return names_by_target


def _iter_xlsx_sheet_members(zf, *, first_sheet=False):
    members = [
        name for name in zf.namelist()
        if name.startswith("xl/worksheets/") and name.endswith(".xml")
    ]
    members.sort(key=lambda name: (_xlsx_sheet_number(name), name))
    if first_sheet:
        members = members[:1]
    sheet_names = _xlsx_sheet_names(zf)
    for name in members:
        yield name, sheet_names.get(name, PurePosixPath(name).stem)


def _row_has_cell_value(row_elem):
    for cell in row_elem:
        if _xml_local_name(cell.tag) != "c":
            continue
        for child in cell:
            if _xml_local_name(child.tag) in {"v", "is", "f"}:
                return True
    return False


def _count_xlsx_sheet_rows(zf, member_name):
    physical_rows = 0
    with zf.open(member_name, "r") as stream:
        for _event, elem in ET.iterparse(stream, events=("end",)):
            if _xml_local_name(elem.tag) != "row":
                continue
            if _row_has_cell_value(elem):
                physical_rows += 1
            elem.clear()
    return physical_rows


def _count_xlsx_zip(path_or_file, logical_path, *, header_rows, first_sheet=False):
    records = []
    with zipfile.ZipFile(path_or_file, "r") as zf:
        for member_name, sheet_name in _iter_xlsx_sheet_members(zf, first_sheet=first_sheet):
            physical_rows = _count_xlsx_sheet_rows(zf, member_name)
            records.append(
                CountRecord(
                    path=logical_path,
                    kind="excel",
                    sheet=sheet_name,
                    physical_rows=physical_rows,
                    data_rows=_data_rows(physical_rows, header_rows),
                    header_rows=min(header_rows, physical_rows),
                )
            )
    return records


def _count_xls_file_with_pandas(path, *, header_rows, first_sheet=False):
    import pandas as pd

    sheet_name = 0 if first_sheet else None
    sheets = pd.read_excel(path, sheet_name=sheet_name, header=None)
    if first_sheet:
        sheets = {"0": sheets}
    records = []
    for sheet, frame in sheets.items():
        physical_rows = int(frame.dropna(how="all").shape[0])
        records.append(
            CountRecord(
                path=str(path),
                kind="excel",
                sheet=str(sheet),
                physical_rows=physical_rows,
                data_rows=_data_rows(physical_rows, header_rows),
                header_rows=min(header_rows, physical_rows),
            )
        )
    return records


def _count_xls_bytes_with_pandas(data, logical_path, *, header_rows, first_sheet=False):
    import pandas as pd

    sheet_name = 0 if first_sheet else None
    sheets = pd.read_excel(BytesIO(data), sheet_name=sheet_name, header=None)
    if first_sheet:
        sheets = {"0": sheets}
    records = []
    for sheet, frame in sheets.items():
        physical_rows = int(frame.dropna(how="all").shape[0])
        records.append(
            CountRecord(
                path=logical_path,
                kind="excel",
                sheet=str(sheet),
                physical_rows=physical_rows,
                data_rows=_data_rows(physical_rows, header_rows),
                header_rows=min(header_rows, physical_rows),
            )
        )
    return records


def _error_record(path, kind, exc):
    return CountRecord(
        path=str(path),
        kind=kind,
        physical_rows=0,
        data_rows=0,
        header_rows=0,
        error=f"{type(exc).__name__}: {exc}",
    )


def _count_table_file(path, *, header_rows, first_sheet=False):
    ext = _suffix(path)
    try:
        if ext in CSV_EXTS:
            return [_count_csv_file(path, header_rows=header_rows)]
        if ext in JSONL_EXTS:
            return [_count_jsonl_file(path)]
        if ext == ".xls":
            return _count_xls_file_with_pandas(path, header_rows=header_rows, first_sheet=first_sheet)
        if ext in EXCEL_EXTS:
            return _count_xlsx_file(path, header_rows=header_rows, first_sheet=first_sheet)
    except Exception as exc:  # Keep batch counting robust.
        return [_error_record(path, "table", exc)]
    return []


def _count_table_bytes(data, logical_path, *, header_rows, first_sheet=False):
    ext = _suffix(logical_path)
    try:
        if ext in CSV_EXTS:
            return [_count_csv_bytes(data, logical_path, header_rows=header_rows)]
        if ext in JSONL_EXTS:
            return [_count_jsonl_bytes(data, logical_path)]
        if ext == ".xls":
            return _count_xls_bytes_with_pandas(
                data,
                logical_path,
                header_rows=header_rows,
                first_sheet=first_sheet,
            )
        if ext in EXCEL_EXTS:
            return _count_xlsx_bytes(
                data,
                logical_path,
                header_rows=header_rows,
                first_sheet=first_sheet,
            )
    except Exception as exc:  # Keep batch counting robust.
        return [_error_record(logical_path, "table", exc)]
    return []


def _iter_zip_records(zip_bytes_or_path, logical_path, *, header_rows, first_sheet=False):
    try:
        zf = zipfile.ZipFile(zip_bytes_or_path, "r")
    except Exception as exc:
        yield _error_record(logical_path, "zip", exc)
        return
    with zf:
        for info in sorted(zf.infolist(), key=lambda item: item.filename):
            if info.is_dir():
                continue
            member_name = info.filename
            member_path = f"{logical_path}!/{member_name}"
            ext = PurePosixPath(member_name).suffix.lower()
            if ext not in (CSV_EXTS | JSONL_EXTS | EXCEL_EXTS | ZIP_EXTS):
                continue
            try:
                data = zf.read(info)
            except Exception as exc:
                yield _error_record(member_path, "zip_member", exc)
                continue
            if ext in ZIP_EXTS:
                yield from _iter_zip_records(
                    BytesIO(data),
                    member_path,
                    header_rows=header_rows,
                    first_sheet=first_sheet,
                )
            else:
                yield from _count_table_bytes(
                    data,
                    member_path,
                    header_rows=header_rows,
                    first_sheet=first_sheet,
                )


def iter_count_records(root, *, header_rows=1, first_sheet=False):
    root_path = Path(root)
    if root_path.is_file():
        ext = _suffix(root_path)
        if ext in ZIP_EXTS:
            yield from _iter_zip_records(
                root_path,
                str(root_path),
                header_rows=header_rows,
                first_sheet=first_sheet,
            )
        elif _is_supported_table(root_path):
            yield from _count_table_file(
                root_path,
                header_rows=header_rows,
                first_sheet=first_sheet,
            )
        return

    for dirpath, dirnames, filenames in os.walk(root_path):
        dirnames.sort()
        for filename in sorted(filenames):
            path = Path(dirpath) / filename
            ext = _suffix(path)
            if ext in ZIP_EXTS:
                yield from _iter_zip_records(
                    path,
                    str(path),
                    header_rows=header_rows,
                    first_sheet=first_sheet,
                )
            elif _is_supported_table(path):
                yield from _count_table_file(
                    path,
                    header_rows=header_rows,
                    first_sheet=first_sheet,
                )


def _print_text_summary(records):
    total_data_rows = sum(record.data_rows for record in records if not record.error)
    total_physical_rows = sum(record.physical_rows for record in records if not record.error)
    error_count = sum(1 for record in records if record.error)
    table_count = sum(1 for record in records if not record.error)

    print(f"表/Sheet 数: {table_count}")
    print(f"原始非空行数: {total_physical_rows}")
    print(f"扣表头后数据行数: {total_data_rows}")
    if error_count:
        print(f"读取失败数: {error_count}")


def _print_details(records):
    print("\n明细:")
    for record in records:
        sheet = f" [{record.sheet}]" if record.sheet else ""
        if record.error:
            print(f"- {record.path}{sheet}: ERROR {record.error}")
            continue
        print(
            f"- {record.path}{sheet}: "
            f"原始非空行={record.physical_rows}, "
            f"表头={record.header_rows}, "
            f"数据行={record.data_rows}"
        )


def _shorten_text(text, max_len=PROGRESS_PATH_MAX_LEN):
    text = str(text)
    if len(text) <= max_len:
        return text
    return "..." + text[-(max_len - 3):]


def _progress_extra(records, current_path=""):
    total_data_rows = sum(record.data_rows for record in records if not record.error)
    error_count = sum(1 for record in records if record.error)
    table_count = sum(1 for record in records if not record.error)
    parts = [
        f"表/Sheet {table_count}",
        f"数据行 {total_data_rows}",
    ]
    if error_count:
        parts.append(f"失败 {error_count}")
    if current_path:
        parts.append(_shorten_text(current_path))
    return "，".join(parts)


def collect_count_records(root, *, header_rows=1, first_sheet=False, show_progress=True):
    records = []
    progress = ProgressBar(0, "统计表格行数") if show_progress else None
    try:
        if progress is not None:
            progress.set_extra_text("正在扫描并统计...", force=True)
        for record in iter_count_records(
            root,
            header_rows=header_rows,
            first_sheet=first_sheet,
        ):
            records.append(record)
            if progress is not None:
                progress.set(len(records))
                progress.set_extra_text(_progress_extra(records, record.path))
        if progress is not None:
            progress.set_extra_text(_progress_extra(records, "完成"), force=True)
    finally:
        if progress is not None:
            progress.close()
    return records


def main():
    parser = ArgumentParser(description="递归统计目录/ZIP 内 CSV、JSONL 和 Excel 的数据行数，CSV/Excel 默认扣除表头")
    parser.add_argument("input", help="输入目录、文件或 zip")
    parser.add_argument(
        "--header-rows",
        type=int,
        default=1,
        help="每个 CSV 文件/Excel sheet 扣除的表头行数，默认 1；JSONL 始终不扣表头",
    )
    parser.add_argument(
        "--first-sheet",
        action="store_true",
        help="Excel 只统计第一个 sheet；默认统计所有 sheet",
    )
    parser.add_argument("--details", action="store_true", help="打印每个文件/sheet 的明细")
    parser.add_argument("--json-output", default="", help="可选：把统计结果写为 JSON")
    parser.add_argument("--no-progress", action="store_true", help="关闭进度显示")
    args = parser.parse_args()

    if args.header_rows < 0:
        raise ValueError("--header-rows must be >= 0")

    records = collect_count_records(
        args.input,
        header_rows=args.header_rows,
        first_sheet=args.first_sheet,
        show_progress=not args.no_progress,
    )
    _print_text_summary(records)
    if args.details:
        _print_details(records)

    if args.json_output:
        payload = {
            "input": str(Path(args.input).resolve()),
            "header_rows": args.header_rows,
            "first_sheet": bool(args.first_sheet),
            "summary": {
                "table_count": sum(1 for record in records if not record.error),
                "physical_rows": sum(record.physical_rows for record in records if not record.error),
                "data_rows": sum(record.data_rows for record in records if not record.error),
                "error_count": sum(1 for record in records if record.error),
            },
            "records": [record.to_dict() for record in records],
        }
        with open(args.json_output, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        print(f"JSON 结果已写入: {args.json_output}")


if __name__ == "__main__":
    main()
